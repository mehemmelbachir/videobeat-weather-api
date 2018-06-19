import requests
from weather.settings import API_URL
from .models import InputData, OutputData
from openpyxl import load_workbook
from datetime import datetime
from django.db.models import Avg


class APIError(Exception):
    """An API Error Exception"""

    def __init__(self, status):
        self.status = status

    def __str__(self):
        return "APIError: status={}".format(self.status)


def get_initial_data():
    try:
        # Getting file from API
        response = requests.get(API_URL)
        if response.status_code != 200:
            raise APIError(response.status_code)

        # Downloading xlsx file
        with open('initial.xlsx','wb') as f:
            f.write(response.content)

        # Loading xlsx file
        workbook = load_workbook('initial.xlsx', read_only=True)
        first_sheet = workbook.sheetnames[0]
        worksheet = workbook[first_sheet]

        # Parsing and saving Input data into our database
        rows = list(worksheet.iter_rows())[1:]
        for row in rows:
            # Create record only if it doesn't exists
            if not InputData.objects.filter(date = datetime.strptime(row[0].value, '%Y-%m-%d'),
                                            time=row[1].value,
                                            latitude=row[2].value,
                                            longitude=row[3].value).exists():

                InputData.objects.create(
                                        date = datetime.strptime(row[0].value,'%Y-%m-%d'),
                                        time = row[1].value,
                                        latitude = row[2].value,
                                        longitude = row[3].value,
                                        hight_ug = row[4].value,
                                        temperature = row[5].value,
                                        wind_speed = row[6].value,
                                        wind_vector_direction = row[7].value)
    except Exception as ex:
        template = "An exception of type {0} occurred. Arguments:\n{1!r}"
        message = template.format(type(ex).__name__, ex.args)
        print(message)


# Algorithme that calculate Output data and store it in database...
def calculate_output_data():
    """
    Just generate average values and store them on the database
    """
    try:
        # Clear data...
        OutputData.objects.all().delete()

        # Calculate new values
        avg_temperature = InputData.objects.all().aggregate(Avg('temperature'))
        avg_wind_speed = InputData.objects.all().aggregate(Avg('wind_speed'))
        avg_direction = InputData.objects.all().aggregate(Avg('wind_vector_direction'))

        OutputData.objects.create(avg_temperature = avg_temperature['temperature__avg'],
                                  avg_wind_speed = avg_wind_speed['wind_speed__avg'],
                                  direction_in_space = avg_direction['wind_vector_direction__avg']
                                  )
    except Exception as ex:
        template = "An exception of type {0} occurred. Arguments:\n{1!r}"
        message = template.format(type(ex).__name__, ex.args)
        print(message)
